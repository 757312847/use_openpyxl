import os
import openpyxl
import re

name1 = "退市前一年"
name2 = "退市前二年"
name3 = "退市前三年"

#常用参数
dir_Path = os.path.dirname(__file__)
out_name = "{}\\{}".format(dir_Path,"merge.xlsx")
file_type = ".xlsx"

#删除文件，避免被选中
os.remove(out_name)

#以后缀名为索引，返回当前文件夹所有文件名-list
def getFiles(dir, suffix): # （根目录，文件后缀 ）
    res = []
    for root, directory, files in os.walk(dir):  # =>当前根,根下目录,目录下的文件
        for filename in files:
            name, suf = os.path.splitext(filename) # =>文件名,文件后缀
            if suf == suffix:
                res.append(filename) # =>吧一串字符串组合成路径
    return res

#提取列表中字符串的数字
def get_digit(list_name):
    re_list = list()
    for i in list_name:
        re_list.append(re.sub("\D", "", i))
    return re_list

filename = getFiles(dir_Path,file_type)
file_num = get_digit(filename)

#创建工作簿和sheet
wb = openpyxl.Workbook()
sheet1 = wb.worksheets[0]
sheet1.title = name1
sheet2 = wb.create_sheet(title=name2)
sheet3 = wb.create_sheet(title=name3)

#创建和处理表头
headers_ws = openpyxl.load_workbook("{}\\{}".format(dir_Path,filename[0])).worksheets[0]
headers = [item.value for item in list(headers_ws.rows)[0]]
del headers[0]
headers.insert(0,"Index")

#写入表头
for i in range(1,len(headers)):
    sheet1.cell(row = 1,column = i,value = headers[i-1])
    sheet2.cell(row = 1,column = i,value = headers[i-1])
    sheet3.cell(row = 1,column = i,value = headers[i-1])

#写入数据
row = 2
for (name,num) in zip(filename,file_num):
    #提取行（不包括表头），处理每一行数据
    data_ws = openpyxl.load_workbook("{}\\{}".format(dir_Path,name)).worksheets[0]
    data1 = [item.value for item in list(data_ws.rows)[1]]
    data2 = [item.value for item in list(data_ws.rows)[2]]
    data3 = [item.value for item in list(data_ws.rows)[3]]
    del data1[0],data2[0],data3[0]
    num = int(num)
    data1.insert(0,num)
    data2.insert(0,num)
    data3.insert(0,num)

    #写入数据
    for i in range(1,len(data1)):
        sheet1.cell(row = row,column = i,value = data1[i-1])
        sheet2.cell(row = row,column = i,value = data2[i-1])
        sheet3.cell(row = row,column = i,value = data3[i-1])
    row += 1
    
wb.save(out_name)


